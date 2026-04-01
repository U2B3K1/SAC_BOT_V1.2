import io
from datetime import date
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.core.database import get_supabase_admin

db = get_supabase_admin()


def _border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor="1E3A5F")


def _subheader_fill():
    return PatternFill("solid", fgColor="2E86C1")


def _total_fill():
    return PatternFill("solid", fgColor="F0F3F4")


def generate_report_excel(
    report_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    department_id: Optional[str] = None,
) -> bytes:
    """Excel hisobot generatsiya"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    # Sarlavha
    ws.merge_cells("A1:H1")
    ws["A1"] = "RESTORAN BOSHQARUV TIZIMI — HISOBOT"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _header_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sana qatori
    ws.merge_cells("A2:H2")
    date_label = f"Hisobot sanasi: {date.today().isoformat()}"
    if date_from and date_to:
        date_label = f"Davr: {date_from} – {date_to}"
    ws["A2"] = date_label
    ws["A2"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = _subheader_fill()
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Ma'lumotlarni olish
    reports = _get_reports(report_id, date_from, date_to, department_id)

    row = 4
    for report in reports:
        dept_name = (report.get("departments") or {}).get("name", "")
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"📅 {report['report_date']}  |  Bo'lim: {dept_name}"
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="2C3E50")
        ws.row_dimensions[row].height = 18
        row += 1

        # Sarlavhalar: Sotuvlar
        headers = ["Mahsulot", "Miqdor", "Narx", "Jami", "Tannarx", "Foyda", "Bo'lim", "Usul"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = _subheader_fill()
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
        ws.row_dimensions[row].height = 16
        row += 1

        # Sotuv satrlari
        sales = db.table("sales").select(
            "*, products(name, departments(name))"
        ).eq("daily_report_id", report["id"]).execute().data

        for sale in sales:
            product_name = (sale.get("products") or {}).get("name", "")
            dept = ((sale.get("products") or {}).get("departments") or {}).get("name", "")
            total = sale.get("total_amount", 0) or 0
            cost = sale.get("total_cost", 0) or 0
            profit = total - cost

            row_data = [
                product_name,
                sale.get("quantity", 0),
                sale.get("unit_price", 0),
                total,
                cost,
                profit,
                dept,
                sale.get("input_method", "manual"),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = _border()
                if col in [3, 4, 5, 6]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        # Xarajatlar bo'limi
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "XARAJATLAR"
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="C0392B")
        row += 1

        exp_headers = ["Kategoriya", "Summa", "Izoh", ""]
        for col, h in enumerate(exp_headers[:3], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = _total_fill()
        row += 1

        expenses = db.table("expenses").select(
            "*, expense_categories(name)"
        ).eq("daily_report_id", report["id"]).execute().data

        for exp in expenses:
            cat_name = (exp.get("expense_categories") or {}).get("name", "")
            ws.cell(row=row, column=1, value=cat_name).border = _border()
            cell = ws.cell(row=row, column=2, value=exp.get("amount", 0))
            cell.border = _border()
            cell.number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=exp.get("description", "")).border = _border()
            row += 1

        # Yig'ma
        row += 1
        summary_data = [
            ("Jami daromad:", report.get("total_revenue", 0)),
            ("Tannarx:", report.get("total_cost", 0)),
            ("Yalpi foyda:", report.get("gross_profit", 0)),
            ("Xarajatlar:", report.get("total_expenses", 0)),
            ("Sof foyda:", report.get("net_profit", 0)),
            ("Balans:", report.get("closing_balance", 0)),
        ]
        for label, value in summary_data:
            ws.cell(row=row, column=6, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=7, value=value)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            row += 1

        row += 2

    # Ustun kengliklari
    col_widths = [30, 10, 14, 14, 14, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Saqlash
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _get_reports(report_id, date_from, date_to, department_id):
    q = db.table("daily_reports").select("*, departments(name)")
    if report_id:
        q = q.eq("id", report_id)
    elif date_from and date_to:
        q = q.gte("report_date", date_from.isoformat()).lte("report_date", date_to.isoformat())
    if department_id:
        q = q.eq("department_id", department_id)
    return q.order("report_date").execute().data
